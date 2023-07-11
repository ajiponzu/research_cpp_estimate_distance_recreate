# 引用: 「https://megatenpa.com/python/base/excel-read-maek-graph/」

import glob
import natsort
import pandas as pd
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series

# 各csvファイルの各データごとにグラフを作成
print(' - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -')
# 読み込みファイルの共通した接頭辞・接尾辞・拡張子
prefix = 'output_'
suffix = ''
extension = '.csv'
# 以下の形式でファイルを読み込む
file_name = f'{prefix}*{suffix}{extension}'
# このpyファイルと同じディレクトリのファイルをソートして抽出
file_list = natsort.natsorted(glob.glob(file_name))
print(file_list)
print(' - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -')

for name in file_list:
    # 出力用のExcelファイル（ワークブック）の作成
    wb = openpyxl.Workbook()
    # 作成したワークブックを保存
    output_name = f'{name}.xlsx'
    wb.save(output_name)
    # 出力用のワークブックの読み込み
    wb = openpyxl.load_workbook(output_name)
    # ワークシートの選択（アクティブなワークシートを自動選択）
    ws = wb.active

    # csvファイルをpandasで読み込み
    df = pd.read_csv(name)
    # 値の部分だけ取り出し
    ws_tmp = df.values
    # ヘッダーの作成
    # .columnsでデータフレームからヘッダーを取り出す
    header_x = f"{df.columns[0]}"
    header_y = f"{df.columns[1]}"
    # 書き込む先の列を決める
    # chrでアルファベットに変換
    cell_x = chr(65)  # 初めはA列、その次はC列
    cell_y = chr(66)  # 初めはB列、その次はD列
    ws[f'{cell_x}1'] = header_x
    ws[f'{cell_y}1'] = header_y
    # データフレームから取り出した値を行ごとに回す
    for num, rows in enumerate(ws_tmp, 1):
        x = rows[0]  # 各行の横軸の値
        y = rows[1]  # 各行の縦軸の値
        # 出力用のワークブックに書き込み
        ws[f'{cell_x}{num + 1}'] = x
        ws[f'{cell_y}{num + 1}'] = y
    # グラフ化するためのScatterChartオブジェクト作成
    chart = ScatterChart()
    chart.x_axis.title = f"{header_x}"
    chart.y_axis.title = f"{header_y} [m]"
    # グラフの作成
    # 最大の行数 = 書き込む最後の行番号
    max_row = len(ws_tmp)
    # xの範囲を設定
    x_values = Reference(
        ws,  # ワークシート
        min_col=1,  # 初めの列番号（1ファイルごとにずらす）
        min_row=2,  # 初めの行番号（ヘッダーを除くので2スタート）
        max_row=max_row + 1  # 最後の行番号
    )
    # yの範囲を設定
    values = Reference(
        ws,
        min_col=2,
        min_row=2,
        max_row=max_row + 1
    )
    # グラフの追加
    series = Series(values, x_values, title=header_y)
    chart.series.append(series)
    # グラフを設置B列から設置（設置位置は微調節）
    position = f"{chr(70)}{10}"  # グラフを設置するセル番地
    ws.add_chart(chart, position)  # グラフの設置
    # ワークブックの保存
    wb.save(output_name)
